import os
from openpyxl import load_workbook

from utils import *
from default_path import template_path
from reader import RecReaderLoop
from writer import PBU_Pressure_writer

class TemplateFiller():
    def __init__(self, target: str = "") -> None:
        self.target = target
        self.template = self.InitializeTemplate()
        self.workbook = load_workbook(self.template)

    def InitializeTemplate(self):
        options = Options(["[!] FBHP", "[!] SBHP", "PBUPDD", "[!] Multirate", "[!] FOT", "[!] RFT/MDT"])
        option = Choose(options)

        if option == 3:
            return os.path.abspath(template_path[3])

    @RecReaderLoop
    @PBU_Pressure_writer
    def Rec_Parser():
        pass #TODO

    def OpenpyxlSandbox(self):
        # Print Worksheet
        for sheet in self.workbook:
            print(sheet.title, end=" | ")
        
        pressure_sheet = self.workbook["PRESSURE"]
        cell = pressure_sheet.cell(4,1)
        cell.value = 10
        self.workbook.save(self.template)