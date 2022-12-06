from typing import Any, List, Tuple

import os
from zipfile import BadZipFile
import xlrd
import openpyxl

from utils import *
from filter import FilterCLI
from xlsx_writer import ExcelWriter

class ExcelReadError(Exception):
    pass

class ExcelReader:
    def __init__(self, path: str) -> None:
        self.path = path
        self.filename = os.path.basename(path)
        self.ext = os.path.splitext(path)[1].lower()
        self.row = 1
        self.col = 1

        self.Init()

    def Init(self) -> None:
        try:
            self.OpenBook()
        # except BadZipFile as e:
        #     raise ExcelReadError(f"Cannot open {self.filename} as {e}.")
        # except xlrd.XLRDError as e:
        #     raise ExcelReadError(f"Cannot open {self.filename} as {e}.")
        # except KeyError as e:
        #     raise ExcelReadError(f"Cannot open {self.filename} as {e}.")
        except Exception as e:
            raise ExcelReadError(f"Cannot open {self.filename} as {e}.")

        self.sheetnames = self.GetSheetnames()
        self.SelectSheet(self.sheetnames[0]) # Select first sheet

    def OpenBook(self) -> None:
        if self.ext == ".xls":
            self.book = xlrd.open_workbook(self.path, on_demand=True)
        elif self.ext in [".xlsx", ".xlsm"]:
            self.book = openpyxl.load_workbook(self.path)

    def SelectSheet(self, sheet_name: str) -> None:
        if self.ext == ".xls":
            self.sheet = self.book.sheet_by_name(sheet_name)
        elif self.ext in [".xlsx", ".xlsm"]:
            self.sheet = self.book[sheet_name]

    def SelectCell(self, row: int, col: int) -> None:
        self.row = row if row > 0 else 1
        self.col = col if col > 0 else 1

    def ReadCell(self, row: int, col: int) -> str:
        if self.ext == ".xls":
            try:
                return str(self.sheet.cell_value(row - 1, col - 1))
            except IndexError:
                return ""

        elif self.ext in [".xlsx", ".xlsm"]:
            value = self.sheet.cell(row, col).value
            return str(value) if value else ""

    def ReadSelectedCell(self) -> Any:
        return self.ReadCell(self.row, self.col)

    def GetSheetnames(self) -> List[str]:
        if self.ext == ".xls":
            return [str(_) for _ in self.book.sheet_names()]
        elif self.ext in [".xlsx", ".xlsm"]:
            return [str(_) for _ in self.book.sheetnames]

    def NextRow(self) -> None:
        self.row = self.row + 1

    def NextCol(self) -> None:
        self.col = self.col + 1

    # f_* means initial position
    # d_* means range of number to find
    def FindCellByKey(self, keys: List[str], f_row = 1, f_col = 1, d_row = 2, d_col = 2) -> Tuple[int, int]:
        for i_row in range(d_row):
            row = f_row + i_row
            for j_col in range(d_col):
                col = f_col + j_col
                if KeysIn(self.ReadCell(row, col).lower(), *keys):
                    return row, col
        return -1, -1

    def FindCellsByKey(self, keys: List[str], f_row = 1, f_col = 1, d_row = 2, d_col = 2) -> List[Tuple[int, int]]:
        row_cols = []
        for i_row in range(d_row):
            row = f_row + i_row
            for j_col in range(d_col):
                col = f_col + j_col
                if KeysIn(self.ReadCell(row, col).lower(), *keys):
                    row_cols.append((row, col))
        return row_cols

    @staticmethod
    def _ColToNum(col: str) -> int:
        if len(col) != 1: raise ValueError("Input only 1 char.")

        alphabets = "abcdefghijklmnopqrstuvwxyz"
        for i in range(len(alphabets)):
            if col.lower() == alphabets[i]:
                return i + 1

        raise ValueError("Column is not in alphabets!")

    @staticmethod
    def ColToNum(col: str) -> int:
        if len(col) == 1:
            return ExcelReader._ColToNum(col)

        return ExcelReader._ColToNum(col[0]) * pow(26, len(col) - 1) + ExcelReader.ColToNum(col[1:])

def RigReportParser(reader: ExcelReader, keyword_remarks: List[str]) -> Tuple[str, List[str]]:
    # Parse Date
    radius = 4
    s_row = 5
    s_col = ExcelReader.ColToNum("v")
    row, col = reader.FindCellByKey(["WORKING DATE"], s_row - radius, s_col - radius, radius * 2, radius * 2)
    if row == -1: return "", [] # Skip if not rig report
    date = reader.ReadCell(row, col + 1)

    # Parse Remarks
    radius = 5
    s_row = 51
    s_col = ExcelReader.ColToNum("g")
    row, col = reader.FindCellByKey(["OPERATION REMARKS"], s_row - radius, s_col - radius, radius * 2, radius * 2)
    if row == -1: return "", [] # Skip if not rig report
    remarks = []
    for i in range(40):
        value = reader.ReadCell(row + 1 + i, col)

        if value and KeysIn(value, *keyword_remarks):
            for add_row in range(4):
                for add_column in range(4):
                    next_value = reader.ReadCell(row + 1 + i + add_row, col + add_column)
                    spaces = " " if next_value else ""
                    value = value + spaces + next_value
            remarks.append(value)

    return date, remarks

def main():
    WELLS_PATH = Abspath("Input well folders: ")
    FIELD_NAME = os.path.basename(SplitPath(WELLS_PATH, 1))
    WELL_FOLDERS = FilterCLI(Folders(WELLS_PATH), os.path.basename)
    WELL_REL_PATH = Relpath("Input subfolder to search: ")

    writer = ExcelWriter("C:\\Users\\Rogan\\Documents\\Injection Summary", os.path.basename(FIELD_NAME) + " RR Injection Summary")
    writer.CreateHeaderColumn(["Well", "Date", "Remarks", "File Name"])
    
    print("--- Start ---")
    for well in WELL_FOLDERS:
        WELL_NAME = os.path.basename(well)
        well_search_path = os.path.join(well, WELL_REL_PATH)
        # print(well_search_path)
        print(f"~~ ## {WELL_NAME} ## ~~")

        xls_files = Files(well_search_path, "xls", recursive=True)
        # print(xls_files)
        # print()

        for xls_file in xls_files:
            print(f"- {os.path.basename(xls_file)}")
            try:
                reader = ExcelReader(xls_file)
            except ExcelReadError as e:
                print(f"[Warning] {e}. Skipping...")
                continue

            for sheet in reader.sheetnames:
                reader.SelectSheet(sheet)
                date, remarks = RigReportParser(reader, ["conduct injectivity test", "injectivity test"])
                for remark in remarks:
                    writer.Write(col=1, value=WELL_NAME)
                    writer.WriteDate(col=2, value=date)
                    writer.Write(col=3, value=remark)
                    writer.Write(col=4, value=xls_file)
                    writer.NextRow()

    writer.Save()

def main2():
    reader = ExcelReader(Abspath("Input: "))
    print(reader.sheetnames)
    reader.SelectSheet(reader.sheetnames[0])
    value = reader.ReadCell(5,reader.ColToNum('w'))
    print(value)

    writer = ExcelWriter("C:\\Users\\Rogan\\Documents", "ex_write")
    writer.WriteDate(1,1,value)
    writer.Save()

def main3():
    reader = ExcelReader(Abspath("Input path: "))
    date, remarks = RigReportParser(reader, ["portacamp"])
    print(date)
    print(remarks)

if __name__ == "__main__":
    while(True):
        main()